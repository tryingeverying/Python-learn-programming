#！ python3
# updateproduce.py 更改Excel文件中的特定行的值，并将之另存到一个新文件中

import openpyxl

# 读取文件
wb = openpyxl.load_workbook(r'F:\Programming\Python\1_Automate_the_boring_stuff_with_python\course_code\excel file\produceSales.xlsx')
sheet = wb.active


# 待检查的内容
price_update = {
    'Gralic':3.07,
    'Celery':1.19,
    'lemon':1.27
}

# 遍历Excel的所有行
for rownum in range(2,sheet.max_row+1): 
    producename = sheet.cell(row = rownum,column=1).value
    if producename in price_update: # 本程序中最妙的一段代码，判断待检查项是否在要更改的字典中，如果在则用字典中该待更改项的value替代原来value
        sheet.cell(row = rownum,column=2).value = price_update[producename]

wb.save(r'F:\Programming\Python\1_Automate_the_boring_stuff_with_python\course_code\excel file\updaeproduceSales.xlsx')





