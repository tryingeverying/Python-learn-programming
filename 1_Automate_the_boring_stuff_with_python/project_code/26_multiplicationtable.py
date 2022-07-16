#! python3
# 生成输入数字NXN的乘法表

import openpyxl, sys
from openpyxl.styles import Font

# num = int(sys.argv[1])
num = int(input("输入数值"))
# print(num)
wb = openpyxl.Workbook() 
sheet = wb['Sheet']
#设置标签字体加粗
boldfont = Font(bold=True)
#设置边框数值
for i in range(1,num+1):
    sheet.cell(row = i+1,column=1).font = boldfont
    sheet.cell(row = i+1,column=1).value = i

for j in range(1,num+1):
    sheet.cell(row = 1,column=j+1).font = boldfont
    sheet.cell(row = 1,column=j+1).value = j


# 设置乘法之后的目标数值
for m in range(2,num+2):
    for n in range(2,num+2):
       
        sheet.cell(row = m,column=n).value = \
        sheet.cell(row = m,column=1).value * sheet.cell(row = 1,column=n).value

wb.save(r'F:\Programming\Python\1_Automate_the_boring_stuff_with_python\course_code\excel file\26_multiplicationtable.xlsx')
print('done')











